#!/usr/bin/env python3
"""Snapshot the two curated xlsx files from Dropbox to JSON in data/.

Purpose: The xlsx workbooks live in the CRFW Archive (Dropbox). We
snapshot them into data/*-snapshot.json so the Node import scripts can
read without needing openpyxl, and so git diffs show meaningful
content changes when the xlsx is updated.

Run: python3 scripts/snapshot-xlsx.py

Emits:
  data/embeds-snapshot.json
  data/articles-snapshot.json
"""
import json
import os
import sys
from pathlib import Path
from openpyxl import load_workbook

HOME = Path.home()
ARCHIVE = HOME / "Library" / "CloudStorage" / "Dropbox" / "CRFW" / "CRFW Archive" / "_Documentation"
EMBEDS_XLSX = ARCHIVE / "CRFW_Media_Embeds.xlsx"
ARTICLES_XLSX = ARCHIVE / "CRFW_Documentation_Articles.xlsx"

REPO = Path(__file__).resolve().parent.parent
DATA_DIR = REPO / "data"
DATA_DIR.mkdir(exist_ok=True)


def sheet_to_records(ws):
    """Convert an openpyxl worksheet to a list of dicts, using row 1 as headers."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    records = []
    for row in rows[1:]:
        # Skip fully-empty rows
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        rec = {}
        for i, v in enumerate(row):
            if i >= len(headers) or not headers[i]:
                continue
            # Normalize dates and None
            if hasattr(v, "isoformat"):
                v = v.isoformat()[:10]  # YYYY-MM-DD
            rec[headers[i]] = v
        records.append(rec)
    return records


def snapshot(xlsx_path, out_path, include_sheets=None):
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    out = {"source": str(xlsx_path.name), "sheets": {}}
    for sheet_name in wb.sheetnames:
        if include_sheets and sheet_name not in include_sheets:
            continue
        ws = wb[sheet_name]
        out["sheets"][sheet_name] = sheet_to_records(ws)
    with open(out_path, "w") as f:
        json.dump(out, f, indent=2, ensure_ascii=False, default=str)
    print(f"  ✓ {out_path.relative_to(REPO)}")
    for sheet_name, records in out["sheets"].items():
        print(f"      · {sheet_name}: {len(records)} records")


def main():
    print("Snapshotting xlsx → JSON")
    if not EMBEDS_XLSX.exists():
        print(f"  ! EMBEDS_XLSX not found: {EMBEDS_XLSX}", file=sys.stderr)
        sys.exit(1)
    if not ARTICLES_XLSX.exists():
        print(f"  ! ARTICLES_XLSX not found: {ARTICLES_XLSX}", file=sys.stderr)
        sys.exit(1)

    # Skip the README sheets (narrative, not data)
    snapshot(EMBEDS_XLSX, DATA_DIR / "embeds-snapshot.json",
             include_sheets=["Bandcamp Releases", "YouTube Videos", "SoundCloud Tracks", "All Media (flat)"])
    snapshot(ARTICLES_XLSX, DATA_DIR / "articles-snapshot.json",
             include_sheets=["Articles & Press", "Biographical Summary", "Associated Projects"])


if __name__ == "__main__":
    main()
